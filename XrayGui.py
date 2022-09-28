import os
from tkinter import *
from tkinter import filedialog
from tkinter import  messagebox
from tkinter import scrolledtext
from openpyxl import load_workbook

# App Display
app = Tk()
app.geometry("920x650")
app.title("Xray Renaming App")

frame = Frame(app)
frame.pack(pady=30, padx=10)

# Button Commands
def source_path():
    source_fld.delete(0, END)
    path = filedialog.askdirectory(initialdir="/home/", title="Source Folder")
    source_fld.insert(0, path)
    if source_fld.get() == "":
        source_fld.insert(0, "/")
    else:
        source_fld.get()

def destination_path():
    destin_fld.delete(0, END)
    path = filedialog.askdirectory(initialdir="/home/", title="Destination Folder")
    destin_fld.insert(0, path)

def file_name():
    # Get Filename
    excel_path = filedialog.askopenfilename(initialdir="/home/", title="Select Excel File", filetypes=(("Excel Workbook", ".xlsx"),))
    filename.insert(0, excel_path)
    filename.config(state="readonly")

# Run and Print Messages
    # Add Path to Sheet
def rename_files():
    run_btn.config(state="disabled")
    excel_path = filename.get()
    if excel_path=="":
        messagebox.showerror(title="Excel File Missing", message="Excel File Not Found. Select Excel File")
        file_name()
        excel_path = filename.get()
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    ext = ".jpg"

    image_list = []
    # access_num = []
    # duplicates = []
    access_no_list = []

    for row in range(2, worksheet.max_row + 1):
        image_name = worksheet.cell(row=row, column=1)
        access_no = worksheet.cell(row=row, column=2)
        image_values = image_name.value
        access_no_values = access_no.value
        image_list.append(image_values)
        access_no_list.append(access_no_values)

    # for i in access_num:
    #     if i not in access_no_list:
    #         access_no_list.append(i)

    # 
    # Find file path
    def locate_file(name, path):
        for dirpath, dirname, filename in os.walk(path):
            if name in filename:
                return os.path.join(dirpath, name)

    # Rename file path
    renamed_files = []
    unchanged_files = []
    unchanged_acc_no = []

    def file_rename():
        if destin_fld.get()=="":
            messagebox.showerror(title="Destination Missing", message="Destination Not Found. Select Destination Path")
            destination_path()
            destin_fld.insert(0, destin_fld.get())

        counter = 0
        for i in range(0, len(image_list)):
            file_path1 = locate_file(str(image_list[i]) + ext, source_fld.get())
            new_names = access_no_list[i]
            old_filepath = str(file_path1)
            new_filepath = f"{destin_fld.get()}/{str(new_names)}{ext}"

            if old_filepath == 'None':
                message_box.insert(INSERT, f"--{image_list[i]} - File Does Not Exist\n")
                unchanged_files.append(image_list[i])
                unchanged_acc_no.append(access_no_list[i])
                continue
            else:
                if os.path.exists(new_filepath):
                    new_filepath = f"{destin_fld.get()}/{str(new_names)}(1){ext}"
                os.rename(old_filepath, new_filepath)
                message_box.insert(INSERT, f"--{image_list[i]} renamed to {access_no_list[i]}\n")
                renamed_files.append(access_no_list[i])

        final_message = f"""{len(renamed_files)} Files Have Been Renamed and Added to {destin_fld.get()}
                        \n{len(unchanged_files)} Files Have Not Been Renamed Because Files Did Not Exist"""
        messagebox.showinfo(title=None, message=final_message)
        refresher.config(state="active")

        # Add Sheet for unchanged files on the Excel file
        new_sheet = "Unrenamed Files"
        num = 0
        if (len(unchanged_files) > 0):
            if new_sheet not in workbook.sheetnames:
                # workbook.remove_sheet("Unrenamed Files")
                workbook.create_sheet(title=new_sheet, index=1)
                ws1 = workbook[new_sheet]
                ws1["A1"] = "#"
                ws1["B1"] = "Images Not Renamed"
                ws1["C1"] = "Accession Number"
                ws1.delete_rows(2, ws1.max_row)
                # for item in unchanged_acc_no:
                for file in unchanged_files:
                    # for item in unchanged_acc_no:
                    # access_num = item
                    num += 1
                    ws1.append([num, file])
                    # , access_num
                workbook.save(excel_path)
            else:
                del workbook[new_sheet]
                # workbook.remove(new_sheet)
                workbook.create_sheet(title=new_sheet, index=1)
                ws1 = workbook[new_sheet]
                ws1["A1"] = "#"
                ws1["B1"] = "Images Not Renamed"
                ws1["C1"] = "Accession Number"
                ws1.delete_rows(2, ws1.max_row)
                for file in unchanged_files:
                    # for item in unchanged_acc_no:
                        # access_num = item
                    num += 1
                    ws1.append([num, file])
                        # , access_num
                workbook.save(excel_path)

    return file_rename()

# Refresh Gui
def refresh():
    source_fld.delete('0', END)
    destin_fld.delete('0', END)
    filename.config(state="normal")
    filename.delete('0', END)
    message_box.delete('1.0', END)
    run_btn.config(state="active")
    refresher.config(state="disabled")

# source
source_lbl = Label(frame, text="Source Folder")
source_lbl.grid(row=1, column=1, sticky="W")
source_fld = Entry(frame, name="source_folder", width=60, bd=2, fg="#063abd")
source_fld.grid(row=1, column=2, pady=10, padx="10px", ipady='2px', ipadx='5px', sticky="W")
source_btn = Button(frame, text="Browse", background="#aba9a4", command=lambda: source_path())
source_btn.grid(row=1, column=3, pady="5px", padx="2px")

# destination
destin_lbl = Label(frame, text="Destination Folder")
destin_lbl.grid(row=2, column=1, sticky="W")
destin_fld = Entry(frame, name="destination_folder", width=60, bd=2, fg="#063abd")
destin_fld.grid(row=2, column=2, padx="10px", ipady='2px', ipadx='5px', sticky="W")
destin_btn = Button(frame, text="Browse", background="#aba9a4", command=lambda: destination_path())
destin_btn.grid(row=2, column=3, pady="5px", padx="2px")

# Excel Filename
filename_lbl = Label(frame, text="Excel File")
filename_lbl.grid(row=3, column=1, sticky="W")
filename = Entry(frame, name="filename", width=60, bd=2, fg="#063abd")
filename.grid(row=3, column=2, padx="10px", ipady="2px", ipadx="5px", sticky="W")
filename_btn = Button(frame, text="Browse", background="#aba9a4", command=lambda: file_name())
filename_btn.grid(row=3, column=3, pady="5px", padx="2px")

# Show Print Messages
message_box = scrolledtext.ScrolledText(frame, height=19, width=60, wrap=WORD)
message_box.grid(row=4, column=2, pady="20px", ipadx="5px", sticky="N")

# Refresh Button
refresher = Button(frame, text="Refresh", background="#aba9a4", command= lambda: refresh())
refresher.grid(row=4, column=3, padx="2px", ipady="5px", ipadx="7px", sticky="E")

# Run Button
run_btn = Button(frame, text="Run", background="#aba9a4", command= lambda: rename_files())
run_btn.grid(row=5, column=3, padx="2px", ipady="5px", ipadx="20px", sticky="SE")

app.mainloop()
